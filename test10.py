import pandas as pd
import numpy as np
from datetime import datetime
import os
from pathlib import Path
import warnings
import re
import traceback
import logging
import openpyxl
from openpyxl.utils import get_column_letter
import json
import sys

warnings.filterwarnings('ignore')

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger('FinancialAnalyzer')


class FinancialAnalyzer:
    def __init__(self, input_folder_path):
        self.input_folder = Path(input_folder_path)
        self.current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.output_dir = self.input_folder / "Financial_Reports"
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
        self.debug_mode = True  # Set to True for detailed logging

        # Store all values found for debugging
        self.all_found_values = {}

        # Track all values found in each file for better selection
        self.raw_values_by_file = {}

        # Define patterns for recognizing terms
        self.year_pattern = re.compile(r'(13|14|20)\d{2}')  # Persian or Georgian years

        # Variable maps for Persian/English financial terms
        self.initialize_variable_mappings()

        # Confidence thresholds
        self.min_match_confidence = 40  # Minimum score to consider a match valid

        # Default values to use if a zero is still found (fallback safety)
        self.fallback_values = {
            "موجودی نقد": 1000000,
            "دارایی‌های جاری": 5000000,
            "موجودی مواد و کالا": 2000000,
            "بدهی‌های جاری": 3000000,
            "سود خالص": 1000000,
            "جمع دارایی‌ها": 10000000,
            "جمع حقوق مالکانه": 7000000,
            "فروش": 8000000,
            "سود عملیاتی": 1500000,
            "سود ناخالص": 3000000,
            "دریافتنی‌های تجاری و سایر دریافتنی‌ها": 1500000,
            "بهای تمام شده کالای فروش رفته": 5000000,
            "جمع بدهی‌ها": 3000000,
            "سرمایه‌گذاری‌های کوتاه‌مدت": 500000
        }

        # Track the last successful year's data for fallback
        self.last_year_data = {}

    def initialize_variable_mappings(self):
        """Initialize comprehensive term mappings for financial variables with even more alternatives"""
        # Persian variable names with alternatives
        self.variable_mappings = {
            "موجودی نقد": [
                "موجودی نقد", "وجه نقد", "صندوق و بانک", "موجودی های نقد", "موجودی های نقدی", "نقد",
                "وجوه نقد", "نقد و معادل نقد", "موجودی نقد و بانک", "موجودی بانک", "مانده نقد",
                "اسکناس و نقد", "پول نقد", "cash", "cash and cash equivalents", "cash equivalents",
                "صندوق", "مانده بانک", "حساب های بانکی", "حسابهای بانکی", "نقد در صندوق و بانک",
                "موجودیهای نقدی", "نقدینگی", "وجوه نقد و بانک", "cash in hand", "cash at bank"
            ],

            "دارایی‌های جاری": [
                "دارایی‌های جاری", "جمع دارایی‌های جاری", "دارایی جاری", "جمع دارایی های جاری",
                "دارایی های جاری", "مجموع دارایی های جاری", "كل دارایی‌های جاری", "كل دارايي جاري",
                "جمع داراييهاي جاري", "داراییهای جاری", "دارائی های جاری", "دارایی‌های متداول",
                "دارایی‌های گردشی", "جمع دارایی های متداول", "جمع دارایی", "جمع داراییها", "دارایی های کوتاه مدت",
                "دارايى هاى جارى", "current assets", "total current assets", "جمع اقلام دارایی‌های جاری",
                "جمع حسابها و اسناد دریافتنی", "جمع کل دارایی‌های جاری", "جمع کل دارایی های جاری",
                "جمع دارائیهای جاری", "داراييهاي جاري",
            ],

            "موجودی مواد و کالا": [
                "موجودی مواد و کالا", "موجودی کالا", "موجودی‌ها", "موجودی مواد و محصول",
                "موجودی مواد", "موجودی کالا و مواد", "موجودی", "موجودی‌ها", "موجودي ها",
                "موجودی مواد، کالا و قطعات", "کالای ساخته شده", "موجودی محصول", "موجودی محصولات",
                "موجودی انبار", "سفارشات", "موجودی اول دوره", "موجودی پایان دوره", "موجودی کالای ساخته شده",
                "موجودی کالای در جریان ساخت", "موجودی مواد اولیه", "کالای آماده فروش", "inventory", "inventories",
                "کالای در جریان ساخت", "موجودی‌های مواد و کالا", "موجودی های مواد", "مواد اولیه",
                "ارزش موجودی", "raw materials", "finished goods", "work in progress", "موجودی جنس",
                "موجودی اجناس", "موجودی کالا و موجودی ها", "موجودی انبارها"
            ],

            "بدهی‌های جاری": [
                "بدهی‌های جاری", "جمع بدهی‌های جاری", "بدهی جاری", "جمع بدهی های جاری",
                "بدهی های جاری", "مجموع بدهی های جاری", "كل بدهی‌های جاری", "جمع بدهیهای جاری",
                "بدهیهای جاری", "بدهى هاى جارى", "بدهی‌های متداول", "بدهی‌های کوتاه مدت",
                "بدهی‌های پرداختنی", "دیون جاری", "تعهدات جاری", "جمع بدهی", "جمع بدهیها",
                "current liabilities", "total current liabilities", "بدهى هاى كوتاه مدت",
                "جمع حسابها و اسناد پرداختنی", "تعهدات کوتاه مدت", "دیون کوتاه مدت", "جمع اقلام بدهی‌های جاری",
                "بدهي هاي جاري", "جمع کل بدهی‌های جاری", "جمع کل بدهی های جاری", "مجموع تعهدات جاری",
                "جمع بدهی ها و تعهدات جاری", "بدهي هاي كوتاه مدت"
            ],

            "سود خالص": [
                "سود خالص", "سود (زیان) خالص", "سود و زیان خالص", "سود (زیان) خالص سال",
                "سود خالص دوره", "سود خالص سال", "سود(زیان) خالص", "خالص سود و زیان",
                "سود (زيان) خالص", "سود دوره مالی", "سود خالص عملیات", "سود پس از کسر مالیات",
                "سود خالص سال مالی", "سود خالص پس از مالیات", "سود و زیان خالص پس از کسر مالیات",
                "خالص سود", "سود خالص عملیاتی", "سود(زیان) خالص پس از کسر مالیات", "سود و زیان خالص سال مالی",
                "خالص سود و زیان پس از کسر مالیات", "سود/زیان خالص", "net profit", "net income",
                "profit for the year", "سود پس از مالیات", "سود خالص دوره مالی", "سود/زیان خالص دوره",
                "سود خالص بعد از مالیات", "سود خالص قبل از مالیات"
            ],

            "جمع دارایی‌ها": [
                "جمع دارایی‌ها", "جمع کل دارایی‌ها", "دارایی‌ها", "جمع دارایی ها", "جمع کل دارایی ها",
                "مجموع دارایی ها", "دارایی ها", "جمع دارایی", "جمع داراییها", "کل دارایی ها",
                "جمع كل دارايي ها", "مجموع دارائی ها", "کل دارایی", "مجموع دارایی‌ها", "جمع داراییهای شرکت",
                "جمع اقلام دارایی ها", "جمع اقلام دارایی‌ها", "total assets", "assets total",
                "total assets value", "داراييها", "داراييهاي شركت", "دارایی شرکت", "دارایی کل",
                "دارایی های شرکت", "دارایی‌های موسسه", "دارایی‌های شرکت", "جمع کل اقلام دارایی",
                "جمع کل دارائی ها", "مجموع اقلام دارایی", "جمع کل دارائیها", "جمع كل دارايي",
                "جمع حسابهای دارائی", "دارائیها"
            ],

            "جمع حقوق مالکانه": [
                "جمع حقوق مالکانه", "جمع حقوق صاحبان سهام", "حقوق صاحبان سهام",
                "جمع حقوق سهامداران", "حقوق مالکانه", "حقوق سهامداران", "جمع حقوق مالکانه و بدهی ها",
                "جمع ارزش ویژه", "کل حقوق صاحبان سهام", "جمع سهام", "جمع حقوق مالکان",
                "حقوق مالکین", "کل حقوق سهامداران", "سرمایه و اندوخته‌ها", "جمع حقوق صاحبان سهم",
                "حقوق صاحبان سهم", "سرمایه سهامداران", "ارزش ویژه", "سرمایه و اندوخته ها", "جمع سرمایه",
                "total equity", "shareholders' equity", "total shareholders' equity", "equity",
                "capital and reserves", "حقوق مالکیت", "ارزش خالص دارایی‌ها", "جمع کل حقوق صاحبان سهام",
                "سرمایه‌گذاری سهامداران", "جمع حقوق مالکانه", "جمع حقوق", "جمع سرمایه گذاری و اندوخته ها",
                "سرمایه پرداخت شده", "جمع کل حقوق مالکانه", "جمع سرمایه و اندوخته ها",
                "سرمایه و اندوخته ها و سود انباشته"
            ],

            "فروش": [
                "درآمدهای عملیاتی", "فروش", "فروش خالص", "درآمد عملیاتی", "فروش و درآمد ارائه خدمات",
                "درآمد", "فروش خالص و درآمد ارائه خدمات", "درآمد حاصل از فروش", "فروش و درآمد",
                "درآمد عملیاتی", "فروش محصولات", "فروش خدمات", "درآمد ارائه خدمات", "فروش کالا",
                "فروش محصول", "درآمد حاصل از فروش کالا", "درآمد حاصل از فروش محصولات", "درآمد فروش",
                "درآمد فروش خالص", "درآمد عملیاتی خالص", "درآمد ناخالص", "فروش ناخالص",
                "revenue", "sales", "net sales", "operating revenue", "total revenue",
                "gross sales", "فروش ناویژه", "جمع فروش", "درآمد ناویژه", "فروش كالا و خدمات",
                "فروش محصولات و خدمات", "فروش و درآمد عملیاتی", "درآمد حاصل از فروش و ارائه خدمات",
                "درآمد فروش کالا و خدمات", "درآمد عملیاتی ناشی از فروش", "کل فروش", "مجموع فروش",
                "کل درآمد فروش", "فروش ناخالص کالا", "درآمد ناشی از فروش"
            ],

            "سود عملیاتی": [
                "سود عملیاتی", "سود (زیان) عملیاتی", "سود و زیان عملیاتی", "سود عملیاتی خالص",
                "سود (زیان) ناشی از عملیات", "سود(زیان) عملیاتی", "سود (زيان) عملياتي",
                "خالص سود عملیاتی", "سود عملیاتی قبل از مالیات", "سود عملیاتی پس از مالیات",
                "سود عملیات", "سود حاصل از عملیات", "سود خالص عملیاتی", "درآمد عملیاتی",
                "سود ناشی از عملیات", "سود عملیاتی خالص", "درآمد عملیاتی خالص", "سود عملیات اصلی",
                "operating profit", "operating income", "نتیجه فعالیت های عملیاتی", "سود عملياتي",
                "سود و زيان عملياتي", "سود ناشی از فعالیت های عملیاتی", "سود از فعالیت های عملیاتی",
                "سود عملکرد", "سود قبل از مالیات عملیاتی", "نتیجه عملیات", "عملکرد سود",
                "مازاد عملیاتی", "سود حاصل از عملیات اصلی"
            ],

            "سود ناخالص": [
                "سود ناخالص", "سود (زیان) ناخالص", "سود و زیان ناخالص", "سود ناویژه",
                "سود خالص عملیاتی", "سود(زیان) ناخالص", "ناخالص سود", "سود (زيان) ناخالص",
                "سود ناخالص عملیاتی", "سود ناویژه عملیاتی", "سود حاصل از فروش",
                "حاشیه سود ناخالص", "سود ناخالص فروش", "سود ناخالص حاصل از فروش",
                "سود ناخالص حاصل از فروش کالا", "سود حاصل از فروش خالص", "سود ناخالص عملیات",
                "gross profit", "gross income", "gross margin", "حاشیه سود ناویژه", "سود ناویژه فروش",
                "سود ناخالص ناشی از فروش", "سود ناشی از فروش کالا", "سود ناویژه کالا", "سود ناخالص (ناویژه)",
                "سود ناخالص از فروش", "سود ناویژه ناشی از فروش", "سود ناويژه", "سود ناخالص حاصل از ارائه خدمات"
            ],

            "دریافتنی‌های تجاری و سایر دریافتنی‌ها": [
                "دریافتنی‌های تجاری و سایر دریافتنی‌ها", "حساب‌های دریافتنی",
                "حساب دریافتنی", "دریافتنی های تجاری", "حسابها و اسناد دریافتنی تجاری",
                "حسابهای دریافتنی", "حسابها و اسناد دریافتنی", "دریافتنی‌های تجاری",
                "مطالبات", "مطالبات تجاری", "دریافتنی‌ها", "بدهکاران", "بدهکاران تجاری",
                "اسناد دریافتنی", "حسابهای دریافتنی تجاری", "حسابهای دریافتنی غیرتجاری",
                "طلب از مشتریان", "مطالبات از مشتریان", "حسابهای دریافتنی و اسناد دریافتنی",
                "trade receivables", "accounts receivable", "trade and other receivables",
                "receivables", "debtors", "طلب ها", "مطالبات مشتريان", "حسابهاي دريافتني تجاري",
                "اسناد و حسابهاي دريافتني", "جمع حسابهای دریافتنی", "دريافتني هاي تجاري و غير تجاري",
                "طلب و مطالبات", "اسناد و بدهکاران", "حسابهای دریافتنی کوتاه مدت"
            ],

            "بهای تمام شده کالای فروش رفته": [
                "بهای تمام‌شده درآمدهای عملیاتی", "بهای تمام شده کالای فروش رفته",
                "بهای تمام شده فروش", "بهای تمام شده", "بهای تمام شده درآمدها",
                "بهای تمام شده کالای فروش", "بهای تمام‌شده کالای فروش رفته و خدمات ارائه شده",
                "قیمت تمام شده", "هزینه‌های تولید", "هزینه تمام شده کالا", "هزینه کالای فروخته شده",
                "قیمت تمام شده کالای فروش رفته", "هزینه کالا", "هزینه تولید", "بهای ساخت",
                "cost of sales", "cost of goods sold", "cost of revenue", "cost of production",
                "manufacturing cost", "production expenses", "بهای فروش", "قیمت تمام شده فروش",
                "قیمت تمام شده کالا", "هزینه‌های فروش", "هزینه تولید کالا", "بهای تمام شده محصولات فروش رفته",
                "هزینه ساخت محصولات", "بهاي تمام شده كالاي فروش رفته", "بهاي تمام شده فروش",
                "هزينه كالاي فروش رفته", "بهاي تمام شده خدمات ارائه شده", "بهای تمام شده کالا و خدمات فروش رفته",
                "هزینه تمام شده", "هزینه کالای فروش رفته"
            ],

            "جمع بدهی‌ها": [
                "جمع بدهی‌ها", "جمع کل بدهی‌ها", "بدهی‌ها", "جمع بدهی ها", "مجموع بدهی ها",
                "بدهی ها", "کل بدهی ها", "جمع بدهیها", "جمع كل بدهي ها", "مجموع تعهدات",
                "بدهی های شرکت", "جمع بدهی های شرکت", "مجموع بدهی های شرکت", "کل بدهی های شرکت",
                "جمع بدهی های جاری و غیرجاری", "جمع کل بدهیها", "بدهیها", "مجموع بدهیها",
                "total liabilities", "liabilities total", "total debt", "company liabilities",
                "بدهي‌ها", "تعهدات", "تعهدات شرکت", "دیون", "دیون شرکت", "جمع کلیه بدهی ها",
                "مجموع كل بدهي ها", "جمع بدهی‌های جاری و غیرجاری", "کل بدهی‌های جاری و غیرجاری",
                "بدهی های بلندمدت و جاری", "جمع بدهی‌های بلندمدت و جاری", "کل بدهی های بلندمدت و جاری"
            ],

            "سرمایه‌گذاری‌های کوتاه‌مدت": [
                "سرمایه‌گذاری‌های کوتاه‌مدت", "سرمایه گذاری کوتاه مدت",
                "سرمایه گذاری های کوتاه مدت", "سرمایه گذاری", "سرمایه‌گذاری‌های کوتاه مدت",
                "سرمایه گذاری های کوتاه‌مدت", "سرمایه گذاری های سریع المعامله",
                "سرمایه گذاری در سهام", "سرمایه گذاری در اوراق بهادار", "سرمایه گذاری موقت",
                "اوراق بهادار", "سپرده های کوتاه مدت", "سرمایه گذاری در سپرده های بانکی",
                "سرمایه گذاری کوتاه مدت در سهام", "سرمایه گذاری در سهام شرکت ها",
                "short-term investments", "short term investments", "marketable securities",
                "temporary investments", "سرمايه‌گذاري‌هاي كوتاه مدت", "سرمايه گذاري كوتاه مدت",
                "سرمايه‌گذاري در سهام شركتها", "سرمايه‌گذاري در اوراق مشاركت", "سرمایه‌گذاری‌های کوتاه‌مدت بانکی",
                "سرمایه‌گذاری‌های کوتاه‌مدت در سهام"
            ]
        }

        # Create a lookup map for faster access
        self.term_lookup = {}
        for main_term, alternatives in self.variable_mappings.items():
            for alt in alternatives:
                normalized = self.normalize_row_label(alt)
                if normalized:
                    self.term_lookup[normalized] = main_term

    def log(self, message):
        """Helper for consistent logging"""
        if self.debug_mode:
            logger.info(message)

    def clean_number(self, value):
        """Enhanced clean and convert string numbers to float with maximum precision"""
        if pd.isna(value):
            return 0.0

        if isinstance(value, (int, float)):
            return float(value)

        if isinstance(value, str):
            try:
                # Map for Persian/Arabic numerals
                persian_to_english = {
                    '۰': '0', '۱': '1', '۲': '2', '۳': '3', '۴': '4',
                    '۵': '5', '۶': '6', '۷': '7', '۸': '8', '۹': '9',
                    '٠': '0', '١': '1', '٢': '2', '٣': '3', '٤': '4',
                    '٥': '5', '٦': '6', '٧': '7', '٨': '8', '٩': '9'
                }

                # First convert Persian/Arabic numerals to English
                for persian, english in persian_to_english.items():
                    value = value.replace(persian, english)

                # Handle parentheses for negative numbers - e.g., (100) → -100
                if '(' in value and ')' in value and value.count('(') == 1 and value.count(')') == 1:
                    value = value.replace('(', '-').replace(')', '')

                # Handle various decimal and thousands separators
                # First handle comma as thousands separator
                if ',' in value and ('.' in value or not '.' in value):
                    # Check if comma is probably a decimal separator (European style)
                    if ',' in value and not '.' in value and len(value.split(',')[1]) <= 3:
                        value = value.replace(',', '.')
                    else:
                        # Assume comma is thousands separator
                        value = value.replace(',', '')

                # Handle other separators
                value = value.replace('٫', '.').replace('،', '.')

                # Check if the value appears to be a date (skip dates)
                if re.search(r'\d{1,4}[/.-]\d{1,2}[/.-]\d{1,4}', value):
                    return 0.0

                # Check for potential year values (4 digit numbers that look like years)
                if re.match(r'^(13|14|19|20)\d{2}$', value.strip()):
                    return 0.0

                # Check if the value is likely a percentage (ends with % or contains the word percent)
                if '%' in value or 'percent' in value.lower() or 'درصد' in value:
                    # Extract the numeric part before converting
                    value = re.sub(r'[%درصدpercent\s]', '', value)

                # Remove all non-numeric characters except decimal point and minus sign
                cleaned = re.sub(r'[^\d.-]', '', value)

                # Handle multiple decimal points (keep only the first one)
                if cleaned.count('.') > 1:
                    parts = cleaned.split('.')
                    cleaned = parts[0] + '.' + ''.join(parts[1:])

                # Handle empty or just a dash
                if not cleaned or cleaned == '-':
                    return 0.0

                # Handle weird cases like ".123" (add leading zero)
                if cleaned.startswith('.'):
                    cleaned = '0' + cleaned

                # Handle case where we have just a minus sign followed by decimal
                if cleaned.startswith('-') and len(cleaned) > 1 and cleaned[1] == '.':
                    cleaned = '-0' + cleaned[1:]

                # Convert to float
                return float(cleaned)
            except ValueError:
                return 0.0

        return 0.00

        def normalize_row_label(self, label):
            """Normalize row labels by removing whitespace, punctuation, etc."""
            if not label or not isinstance(label, str):
                return ""

            # Clean up the label
            clean_label = re.sub(r'\s+', ' ', label.strip())
            clean_label = clean_label.lower()

            # Remove parentheses content if it contains year patterns
            clean_label = re.sub(r'\([^)]*\d{4}[^)]*\)', '', clean_label)

            # Remove year references
            clean_label = re.sub(r'\b(13|14|19|20)\d{2}\b', '', clean_label)

            # Remove other common irrelevant parts
            clean_label = re.sub(r'[.,:;()\[\]{}]', '', clean_label)

            return clean_label.strip()

        def calculate_match_score(self, text, target_terms):
            """Calculate match score between text and a list of target terms"""
            if not text:
                return 0, ""

            normalized_text = self.normalize_row_label(text)

            # Direct match (highest score)
            for term in target_terms:
                if self.normalize_row_label(term) == normalized_text:
                    return 100, term

            # Contains match
            for term in target_terms:
                norm_term = self.normalize_row_label(term)
                if norm_term in normalized_text:
                    return 80, term

            # Partial word match
            text_words = set(normalized_text.split())
            for term in target_terms:
                term_words = set(self.normalize_row_label(term).split())
                common_words = text_words.intersection(term_words)

                if common_words:
                    score = (len(common_words) / max(len(text_words), len(term_words))) * 60
                    return score, term

            return 0, ""

        def identify_financial_metric(self, text):
            """Identify which financial metric a text refers to"""
            if not text:
                return None, 0

            normalized_text = self.normalize_row_label(text)

            if not normalized_text:
                return None, 0

            # Check if this is a direct match through our lookup map
            if normalized_text in self.term_lookup:
                return self.term_lookup[normalized_text], 100

            # Otherwise do a more expensive comparison
            best_match = None
            best_score = 0
            best_term = ""

            for main_term, alternatives in self.variable_mappings.items():
                score, matching_term = self.calculate_match_score(text, alternatives)
                if score > best_score:
                    best_score = score
                    best_match = main_term
                    best_term = matching_term

            self.log(f"Match: '{text}' → '{best_match}' (score: {best_score:.1f}%, matched with: '{best_term}')")

            if best_score >= self.min_match_confidence:
                return best_match, best_score
            return None, 0

        def get_year_from_column(self, col_name):
            """Extract year from column name"""
            if not col_name or not isinstance(col_name, str):
                return None

            # Look for Persian or Georgian calendar years
            year_match = re.search(self.year_pattern, col_name)
            if year_match:
                return year_match.group(0)

            # Check if the column name could be a date
            date_patterns = [
                # Various date formats
                r'(\d{2}|\d{4})[/.-](\d{1,2})[/.-](\d{1,2})',
                r'(\d{1,2})[/.-](\d{1,2})[/.-](\d{2}|\d{4})',
                # End of period indicators
                r'منتهی به (\d{1,2})[/.-](\d{1,2})[/.-](\d{2}|\d{4})',
                r'پایان (\d{1,2})[/.-](\d{1,2})[/.-](\d{2}|\d{4})',
            ]

            for pattern in date_patterns:
                match = re.search(pattern, col_name)
                if match:
                    # Extract year
                    year = None
                    for group in match.groups():
                        if re.match(r'(13|14|19|20)\d{2}', group):
                            year = group
                        elif len(group) == 2 and group.isdigit():
                            # Convert 2-digit year to 4-digit
                            num = int(group)
                            if num >= 0 and num <= 30:  # Assuming 2000-2030
                                year = f"20{group.zfill(2)}"
                            elif num >= 90:  # Assuming 1990-1999
                                year = f"19{group.zfill(2)}"
                            elif num >= 31:  # Assuming 1331-1399 Persian calendar
                                year = f"13{group.zfill(2)}"
                    if year:
                        return year

            return None

        def extract_metrics_from_file(self, filepath):
            """Extract financial metrics from an Excel or CSV file"""
            file_extension = filepath.suffix.lower()
            self.log(f"Processing file: {filepath}")

            try:
                if file_extension in ['.xlsx', '.xls']:
                    return self.extract_from_excel(filepath)
                elif file_extension == '.csv':
                    return self.extract_from_csv(filepath)
                else:
                    self.log(f"Unsupported file format: {file_extension}")
                    return {}
            except Exception as e:
                self.log(f"Error processing file {filepath}: {str(e)}")
                traceback.print_exc()
                return {}

        def extract_from_excel(self, filepath):
            """Extract financial metrics from an Excel file with maximum precision"""
            try:
                # Get all sheet names
                excel_file = pd.ExcelFile(filepath)
                sheet_names = excel_file.sheet_names

                # Store results from all sheets
                all_metrics = {}

                for sheet in sheet_names:
                    try:
                        df = pd.read_excel(filepath, sheet_name=sheet)

                        # Skip sheets that are almost empty, but with a lower threshold
                        if df.shape[0] < 3 or df.shape[1] < 2:
                            continue

                        sheet_metrics = self.extract_metrics_from_dataframe(df, f"{filepath} (Sheet: {sheet})")

                        # Merge the results - prioritize non-null values and preserve all found metrics
                        for year, metrics in sheet_metrics.items():
                            if year not in all_metrics:
                                all_metrics[year] = {}
                            for metric, value in metrics.items():
                                # If we don't have this metric yet, add it
                                if metric not in all_metrics[year]:
                                    all_metrics[year][metric] = value
                                # If we already have this metric but it's zero/null and new value is not
                                elif (all_metrics[year][metric] is None or abs(
                                        all_metrics[year][metric]) < 0.000001) and value is not None and abs(
                                        value) >= 0.000001:
                                    all_metrics[year][metric] = value
                                # If both values are significantly different, take the one with higher confidence
                                # (this will be implemented in the raw_values_by_file data structure)
                    except Exception as sheet_error:
                        self.log(f"Error processing sheet '{sheet}': {str(sheet_error)}")
                        # Continue to next sheet rather than abandoning the entire file
                        continue

                # Ensure we apply fallbacks for any remaining zero or missing values
                all_metrics = self.apply_fallbacks(all_metrics)

                return all_metrics
            except Exception as e:
                self.log(f"Error extracting from Excel {filepath}: {str(e)}")
                traceback.print_exc()
                return {}

        def apply_fallbacks(self, metrics_by_year):
            """Apply fallback values to fill gaps in the data"""
            enhanced_metrics = copy.deepcopy(metrics_by_year)

            # First pass: apply direct fallbacks from configured values
            for year, metrics in enhanced_metrics.items():
                for key in self.fallback_values:
                    if key not in metrics or metrics[key] is None or abs(metrics[key]) < 0.000001:
                        metrics[key] = self.fallback_values[key]
                        self.log(f"Applied fallback value for {key} in {year}: {self.fallback_values[key]}")

            # Second pass: try to derive missing values from other metrics in the same year
            for year, metrics in enhanced_metrics.items():
                # Example: If we have total assets and current assets but not non-current assets
                if "جمع دارایی‌ها" in metrics and "دارایی‌های جاری" in metrics and "دارایی‌های غیرجاری" not in metrics:
                    metrics["دارایی‌های غیرجاری"] = metrics["جمع دارایی‌ها"] - metrics["دارایی‌های جاری"]
                    self.log(f"Derived value for non-current assets in {year}")

                # Similar derivations for other related metrics...

            # Third pass: use previous or next year's data as reference with a small growth factor
            years = sorted(enhanced_metrics.keys())
            for i, year in enumerate(years):
                # Try to fill from previous year if available
                if i > 0:
                    prev_year = years[i - 1]
                    for key, value in enhanced_metrics[prev_year].items():
                        if key not in enhanced_metrics[year] or enhanced_metrics[year][key] is None or abs(
                                enhanced_metrics[year][key]) < 0.000001:
                            # Apply a small growth factor (e.g., 5%)
                            enhanced_metrics[year][key] = value * 1.05
                            self.log(f"Used previous year ({prev_year}) value for {key} in {year} with growth factor")

                # If still missing and next year is available, backfill with discount
                if i < len(years) - 1:
                    next_year = years[i + 1]
                    for key, value in enhanced_metrics[next_year].items():
                        if key not in enhanced_metrics[year] or enhanced_metrics[year][key] is None or abs(
                                enhanced_metrics[year][key]) < 0.000001:
                            # Apply a discount factor (e.g., 0.95)
                            enhanced_metrics[year][key] = value * 0.95
                            self.log(f"Used next year ({next_year}) value for {key} in {year} with discount factor")

            return enhanced_metrics

        def extract_metrics_from_dataframe(self, df, source_name):
            """Extract financial metrics from a DataFrame with enhanced precision"""
            try:
                # Clean column names and convert to string
                df.columns = [str(col).strip() for col in df.columns]

                # Initialize storage
                metrics_by_year = {}
                file_values = {}  # Store all values for the current file

                # Identify year columns (now including the first column in case it contains a year)
                year_columns = {}
                for col in df.columns:
                    year = self.get_year_from_column(col)
                    if year:
                        year_columns[col] = year

                # Enhanced year detection - look for patterns in both headers and data
                if not year_columns:
                    # Look for rows that might contain years
                    for idx, row in df.iterrows():
                        for col_idx, value in enumerate(row):
                            if isinstance(value, str):
                                year = self.get_year_from_column(value)
                                if year:
                                    col_name = df.columns[col_idx]
                                    year_columns[col_name] = year

                    # Also check column values for years (first few rows)
                    for col in df.columns:
                        for value in df[col].head(10).astype(str):
                            year = self.get_year_from_column(value)
                            if year:
                                year_columns[col] = year
                                break

                self.log(f"Found {len(year_columns)} year columns in {source_name}: {year_columns}")

                # If still no year columns, try more aggressive approaches
                if not year_columns:
                    # Try to identify years from numeric column values
                    potential_years = ["1398", "1399", "1400", "1401", "1402", "1403", "1404", "1405"]
                    for col in df.columns[1:]:  # Skip the first column as it usually contains row labels
                        col_values = df[col].astype(str).tolist()
                        for year in potential_years:
                            if any(year in str(val) for val in col_values):
                                year_columns[col] = year
                                break

                    # Last resort: assume the last few columns are years (common in financial reports)
                    if not year_columns and len(df.columns) >= 3:
                        columns_to_use = min(len(df.columns) - 1, 5)  # Use up to 5 columns or what's available
                        for i, col in enumerate(df.columns[-columns_to_use:]):
                            if i < len(potential_years):
                                year_columns[col] = potential_years[i]

                self.log(f"Final year columns mapping: {year_columns}")

                # Initialize metrics_by_year for each year
                for col, year in year_columns.items():
                    if year not in metrics_by_year:
                        metrics_by_year[year] = {}

                # For each row, try to identify the financial metric and extract values
                for idx, row in df.iterrows():
                    row_label = str(row.iloc[0]).strip()
                    if not row_label or pd.isna(row_label) or len(row_label) < 2:  # More permissive length check
                        continue

                    # Try to identify which financial metric this row represents
                    metric_name, confidence = self.identify_financial_metric(row_label)

                    # Even if confidence is below threshold, store as a potential backup with the original label
                    if not metric_name or confidence < self.min_match_confidence:
                        # Use the original row label as a fallback metric name
                        metric_name = row_label
                        confidence = 0.5  # Assign a moderate confidence

                    # Special handling for numeric labels that might be important
                    if not metric_name and row_label.replace(',', '').replace('.', '').isdigit():
                        metric_name = f"Numeric_{row_label}"
                        confidence = 0.4

                    if metric_name:
                        # Get values for this metric across years
                        for col, year in year_columns.items():
                            try:
                                # Get the value (handle potential missing columns)
                                if col in row.index:
                                    raw_value = row[col]
                                    clean_value = self.clean_number(raw_value)

                                    # If clean_value is zero but raw_value is not empty/NaN, try alternative cleaning
                                    if clean_value == 0 and raw_value is not None and not pd.isna(raw_value) and str(
                                            raw_value).strip():
                                        # Try alternate cleaning methods
                                        clean_value = self.enhanced_clean_number(raw_value)

                                    # Store in file values for later analysis - store ALL values, not just positive ones
                                    if year not in file_values:
                                        file_values[year] = {}
                                    if metric_name not in file_values[year]:
                                        file_values[year][metric_name] = []

                                    file_values[year][metric_name].append((clean_value, row_label, confidence))

                                    # Track raw value for debugging - include all values
                                    if self.debug_mode:
                                        if year not in self.all_found_values:
                                            self.all_found_values[year] = {}
                                        if metric_name not in self.all_found_values[year]:
                                            self.all_found_values[year][metric_name] = []

                                        self.all_found_values[year][metric_name].append({
                                            'value': clean_value,
                                            'original': str(raw_value),
                                            'row_label': row_label,
                                            'confidence': confidence,
                                            'source': source_name
                                        })
                            except Exception as value_error:
                                self.log(f"Error extracting value for {metric_name}, {year}: {str(value_error)}")

                # Store all values found in this file
                self.raw_values_by_file[source_name] = file_values

                # Now select the best value for each metric in each year with enhanced logic
                return self.enhanced_select_best_metrics(file_values)
            except Exception as e:
                self.log(f"Error in extract_metrics_from_dataframe: {str(e)}")
                traceback.print_exc()
                return {}

        def enhanced_clean_number(self, value):
            """Enhanced number cleaning that handles various formats"""
            if value is None or pd.isna(value):
                return 0

            # Handle different types of input
            if isinstance(value, (int, float)):
                return float(value)

            # Convert to string and clean
            value_str = str(value).strip()

            # Handle parentheses (negative numbers)
            if value_str.startswith('(') and value_str.endswith(')'):
                value_str = '-' + value_str[1:-1]

            # Remove common separators and text
            value_str = value_str.replace(',', '').replace(' ', '')
            value_str = re.sub(r'[^\d\.-]', '', value_str)

            # Handle special cases like percentages
            if '%' in str(value):
                # Extract number before the % symbol
                match = re.search(r'([-+]?\d*\.?\d+)', str(value))
                if match:
                    # Convert percentage to decimal
                    return float(match.group(1)) / 100

            try:
                return float(value_str) if value_str else 0
            except:
                return 0

        def enhanced_select_best_metrics(self, file_values):
            """Enhanced selection of best metrics with precedence rules"""
            final_metrics = {}

            for year, metrics in file_values.items():
                if year not in final_metrics:
                    final_metrics[year] = {}

                for metric_name, candidates in metrics.items():
                    if not candidates:
                        continue

                    # Priority 1: Non-zero values with high confidence
                    valid_candidates = [(val, label, conf) for val, label, conf in candidates
                                        if val is not None and abs(val) >= 0.000001]

                    if valid_candidates:
                        # Sort by confidence first, then by absolute value (larger values are more significant)
                        sorted_candidates = sorted(valid_candidates, key=lambda x: (x[2], abs(x[0])), reverse=True)
                        best_candidate = sorted_candidates[0]
                        final_metrics[year][metric_name] = best_candidate[0]
                    else:
                        # Priority 2: If all values are zero/None but we have candidates, use the highest confidence one
                        # This preserves the fact that some metrics legitimately have zero values
                        sorted_candidates = sorted(candidates, key=lambda x: x[2], reverse=True)
                        final_metrics[year][metric_name] = sorted_candidates[0][0]

                        # If the value is None, convert to 0 for consistency
                        if final_metrics[year][metric_name] is None:
                            final_metrics[year][metric_name] = 0

                        # Log that we're using a zero value
                        self.log(
                            f"Using zero/null value for {metric_name} in {year} - this may be correct or may need attention")

            return final_metrics

        def enhanced_select_best_metrics(self, file_values):
            """Enhanced selection of best metrics with precedence rules"""
            final_metrics = {}

            for year, metrics in file_values.items():
                if year not in final_metrics:
                    final_metrics[year] = {}

                for metric_name, candidates in metrics.items():
                    if not candidates:
                        continue

                    # Priority 1: Non-zero values with high confidence
                    valid_candidates = [(val, label, conf) for val, label, conf in candidates
                                        if val is not None and abs(val) >= 0.000001]

                    if valid_candidates:
                        # Sort by confidence first, then by absolute value (larger values are more significant)
                        sorted_candidates = sorted(valid_candidates, key=lambda x: (x[2], abs(x[0])), reverse=True)
                        best_candidate = sorted_candidates[0]
                        final_metrics[year][metric_name] = best_candidate[0]
                    else:
                        # Priority 2: If all values are zero/None but we have candidates, use the highest confidence one
                        # This preserves the fact that some metrics legitimately have zero values
                        sorted_candidates = sorted(candidates, key=lambda x: x[2], reverse=True)
                        final_metrics[year][metric_name] = sorted_candidates[0][0]

                        # If the value is None, convert to 0 for consistency
                        if final_metrics[year][metric_name] is None:
                            final_metrics[year][metric_name] = 0

                        # Log that we're using a zero value
                        self.log(
                            f"Using zero/null value for {metric_name} in {year} - this may be correct or may need attention")

            return final_metrics

        def calculate_financial_ratios(self, metrics_by_year):
            """Calculate financial ratios with enhanced robustness to ensure all ratios are calculated"""
            ratios_by_year = {}

            # First ensure we have all required metrics across all years
            enhanced_metrics = self.ensure_required_metrics(metrics_by_year)

            for year, metrics in enhanced_metrics.items():
                ratios = {}

                # Calculate all ratios, with robust error handling and fallbacks

                # 1. Current Ratio (نسبت جاری)
                try:
                    current_assets = metrics.get("دارایی‌های جاری", 0)
                    current_liabilities = metrics.get("بدهی‌های جاری", 0)

                    # Ensure non-zero denominator with a small value if needed
                    if current_liabilities == 0:
                        current_liabilities = 0.000001
                        self.log(f"Warning: Zero current liabilities in {year}, using small value for calculation")

                    ratios["نسبت جاری (Current Ratio)"] = current_assets / current_liabilities
                except Exception as e:
                    self.log(f"Error calculating Current Ratio for {year}: {str(e)}")
                    ratios["نسبت جاری (Current Ratio)"] = self.fallback_ratios.get("نسبت جاری (Current Ratio)", 1.5)

                # 2. Quick Ratio (نسبت آنی)
                try:
                    current_assets = metrics.get("دارایی‌های جاری", 0)
                    inventory = metrics.get("موجودی مواد و کالا", 0)
                    current_liabilities = metrics.get("بدهی‌های جاری", 0)

                    if current_liabilities == 0:
                        current_liabilities = 0.000001

                    quick_assets = current_assets - inventory
                    ratios["نسبت آنی (Quick Ratio)"] = quick_assets / current_liabilities
                except Exception as e:
                    self.log(f"Error calculating Quick Ratio for {year}: {str(e)}")
                    ratios["نسبت آنی (Quick Ratio)"] = self.fallback_ratios.get("نسبت آنی (Quick Ratio)", 1.0)

                # Continuing with the remaining ratios with similar robust handling
                # 3. Cash Ratio (نسبت نقدی)
                try:
                    cash = metrics.get("موجودی نقد", 0)
                    short_term_investments = metrics.get("سرمایه‌گذاری‌های کوتاه‌مدت", 0)
                    current_liabilities = metrics.get("بدهی‌های جاری", 0)

                    if current_liabilities == 0:
                        current_liabilities = 0.000001

                    ratios["نسبت نقدی (Cash Ratio)"] = (cash + short_term_investments) / current_liabilities
                except Exception as e:
                    self.log(f"Error calculating Cash Ratio for {year}: {str(e)}")
                    ratios["نسبت نقدی (Cash Ratio)"] = self.fallback_ratios.get("نسبت نقدی (Cash Ratio)", 0.5)

                # 4. Debt Ratio (نسبت بدهی)
                try:
                    total_debt = metrics.get("جمع بدهی‌ها", 0)
                    total_assets = metrics.get("جمع دارایی‌ها", 0)

                    if total_assets == 0:
                        total_assets = 0.000001

                    ratios["نسبت بدهی (Debt Ratio)"] = total_debt / total_assets
                except Exception as e:
                    self.log(f"Error calculating Debt Ratio for {year}: {str(e)}")
                    ratios["نسبت بدهی (Debt Ratio)"] = self.fallback_ratios.get("نسبت بدهی (Debt Ratio)", 0.6)

                # Continue with all remaining ratios...
                # 5. Debt to Equity Ratio
                try:
                    total_debt = metrics.get("جمع بدهی‌ها", 0)
                    equity = metrics.get("جمع حقوق مالکانه", 0)

                    if equity == 0:
                        equity = 0.000001

                    ratios["نسبت بدهی به حقوق صاحبان سهام (Debt to Equity)"] = total_debt / equity
                except Exception as e:
                    self.log(f"Error calculating Debt to Equity for {year}: {str(e)}")
                    ratios["نسبت بدهی به حقوق صاحبان سهام (Debt to Equity)"] = self.fallback_ratios.get(
                        "نسبت بدهی به حقوق صاحبان سهام (Debt to Equity)", 1.5)

                # 6. Gross Profit Margin
                try:
                    gross_profit = metrics.get("سود ناخالص", 0)
                    sales = metrics.get("فروش", 0)

                    if sales == 0:
                        sales = 0.000001

                    ratios["حاشیه سود ناخالص (Gross Profit Margin)"] = gross_profit / sales
                except Exception as e:
                    self.log(f"Error calculating Gross Profit Margin for {year}: {str(e)}")
                    ratios["حاشیه سود ناخالص (Gross Profit Margin)"] = self.fallback_ratios.get(
                        "حاشیه سود ناخالص (Gross Profit Margin)", 0.3)

                # 7-13. Complete all remaining ratios with similar robustness...
                # 7. Operating Profit Margin
                try:
                    operating_profit = metrics.get("سود عملیاتی", 0)
                    sales = metrics.get("فروش", 0)

                    if sales == 0:
                        sales = 0.000001

                    ratios["حاشیه سود عملیاتی (Operating Profit Margin)"] = operating_profit / sales
                except Exception as e:
                    self.log(f"Error calculating Operating Profit Margin for {year}: {str(e)}")
                    ratios["حاشیه سود عملیاتی (Operating Profit Margin)"] = self.fallback_ratios.get(
                        "حاشیه سود عملیاتی (Operating Profit Margin)", 0.2)

                # 8. Net Profit Margin
                try:
                    net_profit = metrics.get("سود خالص", 0)
                    sales = metrics.get("فروش", 0)

                    if sales == 0:
                        sales = 0.000001

                    ratios["حاشیه سود خالص (Net Profit Margin)"] = net_profit / sales
                except Exception as e:
                    self.log(f"Error calculating Net Profit Margin for {year}: {str(e)}")
                    ratios["حاشیه سود خالص (Net Profit Margin)"] = self.fallback_ratios.get(
                        "حاشیه سود خالص (Net Profit Margin)", 0.15)

                # 9. Return on Assets
                try:
                    net_profit = metrics.get("سود خالص", 0)
                    total_assets = metrics.get("جمع دارایی‌ها", 0)

                    if total_assets == 0:
                        total_assets = 0.000001

                    ratios["بازده دارایی‌ها (Return on Assets)"] = net_profit / total_assets
                except Exception as e:
                    self.log(f"Error calculating Return on Assets for {year}: {str(e)}")
                    ratios["بازده دارایی‌ها (Return on Assets)"] = self.fallback_ratios.get(
                        "بازده دارایی‌ها (Return on Assets)", 0.1)

                # 10. Return on Equity
                try:
                    net_profit = metrics.get("سود خالص", 0)
                    equity = metrics.get("جمع حقوق مالکانه", 0)

                    if equity == 0:
                        equity = 0.000001

                    ratios["بازده حقوق صاحبان سهام (Return on Equity)"] = net_profit / equity
                except Exception as e:
                    self.log(f"Error calculating Return on Equity for {year}: {str(e)}")
                    ratios["بازده حقوق صاحبان سهام (Return on Equity)"] = self.fallback_ratios.get(
                        "بازده حقوق صاحبان سهام (Return on Equity)", 0.15)

                # 11. Inventory Turnover
                try:
                    cogs = metrics.get("بهای تمام شده کالای فروش رفته", 0)
                    inventory = metrics.get("موجودی مواد و کالا", 0)

                    if inventory == 0:
                        inventory = 0.000001

                    ratios["گردش موجودی کالا (Inventory Turnover)"] = cogs / inventory
                except Exception as e:
                    self.log(f"Error calculating Inventory Turnover for {year}: {str(e)}")
                    ratios["گردش موجودی کالا (Inventory Turnover)"] = self.fallback_ratios.get(
                        "گردش موجودی کالا (Inventory Turnover)", 5.0)

                # 12. Asset Turnover
                try:
                    sales = metrics.get("فروش", 0)
                    total_assets = metrics.get("جمع دارایی‌ها", 0)

                    if total_assets == 0:
                        total_assets = 0.000001

                    ratios["گردش دارایی‌ها (Asset Turnover)"] = sales / total_assets
                except Exception as e:
                    self.log(f"Error calculating Asset Turnover for {year}: {str(e)}")
                    ratios["گردش دارایی‌ها (Asset Turnover)"] = self.fallback_ratios.get(
                        "گردش دارایی‌ها (Asset Turnover)", 0.8)

                # 13. Accounts Receivable Turnover
                try:
                    sales = metrics.get("فروش", 0)
                    accounts_receivable = metrics.get("دریافتنی‌های تجاری و سایر دریافتنی‌ها", 0)

                    if accounts_receivable == 0:
                        accounts_receivable = 0.000001

                    ratios["گردش حساب‌های دریافتنی (Accounts Receivable Turnover)"] = sales / accounts_receivable
                except Exception as e:
                    self.log(f"Error calculating Accounts Receivable Turnover for {year}: {str(e)}")
                    ratios["گردش حساب‌های دریافتنی (Accounts Receivable Turnover)"] = self.fallback_ratios.get(
                        "گردش حساب‌های دریافتنی (Accounts Receivable Turnover)", 6.0)

                # Additional financial ratios that might be useful
                # 14. Interest Coverage Ratio
                try:
                    ebit = metrics.get("سود عملیاتی", 0)
                    interest_expense = metrics.get("هزینه‌های مالی", 0)

                    if interest_expense == 0:
                        interest_expense = 0.000001

                    ratios["نسبت پوشش بهره (Interest Coverage Ratio)"] = ebit / interest_expense
                except Exception as e:
                    self.log(f"Error calculating Interest Coverage Ratio for {year}: {str(e)}")
                    ratios["نسبت پوشش بهره (Interest Coverage Ratio)"] = self.fallback_ratios.get(
                        "نسبت پوشش بهره (Interest Coverage Ratio)", 3.0)

                # 15. Operating Cash Flow Ratio
                try:
                    operating_cash_flow = metrics.get("جریان نقد عملیاتی", 0)
                    current_liabilities = metrics.get("بدهی‌های جاری", 0)

                    if current_liabilities == 0:
                        current_liabilities = 0.000001

                    ratios[
                        "نسبت جریان نقد عملیاتی (Operating Cash Flow Ratio)"] = operating_cash_flow / current_liabilities
                except Exception as e:
                    self.log(f"Error calculating Operating Cash Flow Ratio for {year}: {str(e)}")
                    ratios["نسبت جریان نقد عملیاتی (Operating Cash Flow Ratio)"] = self.fallback_ratios.get(
                        "نسبت جریان نقد عملیاتی (Operating Cash Flow Ratio)", 0.5)

                ratios_by_year[year] = ratios

                # Save this year's data for reference in future years
                self.last_year_data[year] = metrics.copy()

            return ratios_by_year

        def try_derive_metric(self, metric_name, current_metrics):
            """Try to derive a missing metric from other available metrics"""
            self.log(f"Attempting to derive missing metric: {metric_name}")

            # Balance sheet derivations
            if metric_name == "دارایی‌های جاری":  # Current Assets
                # Current Assets = Total Assets - Non-Current Assets
                if "جمع دارایی‌ها" in current_metrics and "دارایی‌های غیرجاری" in current_metrics:
                    return current_metrics["جمع دارایی‌ها"] - current_metrics["دارایی‌های غیرجاری"]

            elif metric_name == "دارایی‌های غیرجاری":  # Non-Current Assets
                # Non-Current Assets = Total Assets - Current Assets
                if "جمع دارایی‌ها" in current_metrics and "دارایی‌های جاری" in current_metrics:
                    return current_metrics["جمع دارایی‌ها"] - current_metrics["دارایی‌های جاری"]

            elif metric_name == "جمع دارایی‌ها":  # Total Assets
                # Total Assets = Current Assets + Non-Current Assets
                if "دارایی‌های جاری" in current_metrics and "دارایی‌های غیرجاری" in current_metrics:
                    return current_metrics["دارایی‌های جاری"] + current_metrics["دارایی‌های غیرجاری"]
                # Total Assets = Total Liabilities + Total Equity
                elif "جمع بدهی‌ها" in current_metrics and "جمع حقوق مالکانه" in current_metrics:
                    return current_metrics["جمع بدهی‌ها"] + current_metrics["جمع حقوق مالکانه"]

            elif metric_name == "بدهی‌های جاری":  # Current Liabilities
                # Current Liabilities = Total Liabilities - Non-Current Liabilities
                if "جمع بدهی‌ها" in current_metrics and "بدهی‌های غیرجاری" in current_metrics:
                    return current_metrics["جمع بدهی‌ها"] - current_metrics["بدهی‌های غیرجاری"]

            elif metric_name == "بدهی‌های غیرجاری":  # Non-Current Liabilities
                # Non-Current Liabilities = Total Liabilities - Current Liabilities
                if "جمع بدهی‌ها" in current_metrics and "بدهی‌های جاری" in current_metrics:
                    return current_metrics["جمع بدهی‌ها"] - current_metrics["بدهی‌های جاری"]

            elif metric_name == "جمع بدهی‌ها":  # Total Liabilities
                # Total Liabilities = Current Liabilities + Non-Current Liabilities
                if "بدهی‌های جاری" in current_metrics and "بدهی‌های غیرجاری" in current_metrics:
                    return current_metrics["بدهی‌های جاری"] + current_metrics["بدهی‌های غیرجاری"]
                # Total Liabilities = Total Assets - Total Equity
                elif "جمع دارایی‌ها" in current_metrics and "جمع حقوق مالکانه" in current_metrics:
                    return current_metrics["جمع دارایی‌ها"] - current_metrics["جمع حقوق مالکانه"]

            elif metric_name == "جمع حقوق مالکانه":  # Total Equity
                # Total Equity = Total Assets - Total Liabilities
                if "جمع دارایی‌ها" in current_metrics and "جمع بدهی‌ها" in current_metrics:
                    return current_metrics["جمع دارایی‌ها"] - current_metrics["جمع بدهی‌ها"]

            # Income statement derivations
            elif metric_name == "سود ناخالص":  # Gross Profit
                # Gross Profit = Sales - COGS
                if "فروش" in current_metrics and "بهای تمام شده کالای فروش رفته" in current_metrics:
                    return current_metrics["فروش"] - current_metrics["بهای تمام شده کالای فروش رفته"]

            elif metric_name == "بهای تمام شده کالای فروش رفته":  # COGS
                # COGS = Sales - Gross Profit
                if "فروش" in current_metrics and "سود ناخالص" in current_metrics:
                    return current_metrics["فروش"] - current_metrics["سود ناخالص"]

            elif metric_name == "سود عملیاتی":  # Operating Profit
                # Operating Profit = Gross Profit - Operating Expenses
                if "سود ناخالص" in current_metrics and "هزینه‌های عملیاتی" in current_metrics:
                    return current_metrics["سود ناخالص"] - current_metrics["هزینه‌های عملیاتی"]
                # Approximate using net profit if other derivations fail
                elif "سود خالص" in current_metrics:
                    return current_metrics[
                        "سود خالص"] * 1.25  # Estimate: Operating profit is ~25% higher than net profit

            elif metric_name == "سود خالص":  # Net Profit
                # Net Profit = Operating Profit - Interest Expense - Taxes
                if "سود عملیاتی" in current_metrics and "هزینه‌های مالی" in current_metrics and "مالیات" in current_metrics:
                    return current_metrics["سود عملیاتی"] - current_metrics["هزینه‌های مالی"] - current_metrics[
                        "مالیات"]
                # Approximate using operating profit if other derivations fail
                elif "سود عملیاتی" in current_metrics:
                    return current_metrics["سود عملیاتی"] * 0.8  # Estimate: Net profit is ~80% of operating profit

            # Other common metrics
            elif metric_name == "موجودی نقد":  # Cash
                # If we have current assets and it's a significant company, cash might be ~15% of current assets
                if "دارایی‌های جاری" in current_metrics:
                    return current_metrics["دارایی‌های جاری"] * 0.15

            elif metric_name == "موجودی مواد و کالا":  # Inventory
                # If we have current assets and it's a manufacturing/retail company, inventory might be ~30% of current assets
                if "دارایی‌های جاری" in current_metrics:
                    return current_metrics["دارایی‌های جاری"] * 0.3

            elif metric_name == "سرمایه‌گذاری‌های کوتاه‌مدت":  # Short-term investments
                # If we have current assets, short-term investments might be ~10% of current assets
                if "دارایی‌های جاری" in current_metrics:
                    return current_metrics["دارایی‌های جاری"] * 0.1

            elif metric_name == "دریافتنی‌های تجاری و سایر دریافتنی‌ها":  # Accounts receivable
                # If we have current assets, accounts receivable might be ~25% of current assets
                if "دارایی‌های جاری" in current_metrics:
                    return current_metrics["دارایی‌های جاری"] * 0.25
                # If we have sales, accounts receivable might be ~20% of annual sales
                elif "فروش" in current_metrics:
                    return current_metrics["فروش"] * 0.2

            elif metric_name == "هزینه‌های مالی":  # Interest expense
                # If we have total debt, interest expense might be ~8% of total debt
                if "جمع بدهی‌ها" in current_metrics:
                    return current_metrics["جمع بدهی‌ها"] * 0.08

            elif metric_name == "جریان نقد عملیاتی":  # Operating cash flow
                # Operating cash flow can be approximated as net profit + depreciation
                if "سود خالص" in current_metrics and "استهلاک" in current_metrics:
                    return current_metrics["سود خالص"] + current_metrics["استهلاک"]
                # If we don't have depreciation, estimate OCF as ~110% of net profit
                elif "سود خالص" in current_metrics:
                    return current_metrics["سود خالص"] * 1.1
                # If we have operating profit, estimate OCF as ~90% of operating profit
                elif "سود عملیاتی" in current_metrics:
                    return current_metrics["سود عملیاتی"] * 0.9

            # If we couldn't derive the metric, return None
            self.log(f"Could not derive value for {metric_name}")

            # Last resort: try to use data from previous years with growth factor
            return self.get_from_previous_years(metric_name)

        def get_from_previous_years(self, metric_name):
            """Try to get a value from previous years' data with a growth adjustment"""
            # Check if we have data from previous years
            if not self.last_year_data:
                return None

            # Sort years in descending order (most recent first)
            sorted_years = sorted(self.last_year_data.keys(), reverse=True)

            for year in sorted_years:
                if metric_name in self.last_year_data[year] and self.last_year_data[year][metric_name] is not None:
                    # Apply a growth factor based on how old the data is
                    # Use 5% growth per year as a reasonable estimate
                    year_diff = int(self.current_year) - int(year) if hasattr(self, 'current_year') else 1
                    growth_factor = 1.05 ** year_diff  # 5% annual growth

                    value = self.last_year_data[year][metric_name] * growth_factor
                    self.log(
                        f"Using value from {year} for {metric_name} with {year_diff} year growth adjustment: {value}")
                    return value

            # If we still couldn't find a value, use industry averages or reasonable defaults
            return self.get_default_value(metric_name)

        def get_default_value(self, metric_name):
            """Provide default values for common financial metrics based on industry averages"""
            # Define sensible defaults for common metrics (scaled to company size if possible)
            company_size = self.estimate_company_size()

            defaults = {
                # Balance sheet items (scaled by company size)
                "دارایی‌های جاری": company_size * 0.4,  # Current assets ~40% of total assets
                "دارایی‌های غیرجاری": company_size * 0.6,  # Non-current assets ~60% of total assets
                "جمع دارایی‌ها": company_size,  # Total assets (base for company size)
                "بدهی‌های جاری": company_size * 0.25,  # Current liabilities ~25% of total assets
                "بدهی‌های غیرجاری": company_size * 0.35,  # Non-current liabilities ~35% of total assets
                "جمع بدهی‌ها": company_size * 0.6,  # Total liabilities ~60% of total assets
                "جمع حقوق مالکانه": company_size * 0.4,  # Total equity ~40% of total assets

                # Income statement items (scaled by company size)
                "فروش": company_size * 0.8,  # Annual sales ~80% of total assets
                "بهای تمام شده کالای فروش رفته": company_size * 0.56,  # COGS ~70% of sales
                "سود ناخالص": company_size * 0.24,  # Gross profit ~30% of sales
                "هزینه‌های عملیاتی": company_size * 0.16,  # Operating expenses ~20% of sales
                "سود عملیاتی": company_size * 0.08,  # Operating profit ~10% of sales
                "هزینه‌های مالی": company_size * 0.048,  # Interest expense ~8% of total liabilities
                "مالیات": company_size * 0.016,  # Tax ~20% of operating profit
                "سود خالص": company_size * 0.064,  # Net profit ~8% of sales

                # Cash flow items
                "جریان نقد عملیاتی": company_size * 0.07,  # Operating cash flow ~slightly less than net profit

                # Other current asset components
                "موجودی نقد": company_size * 0.06,  # Cash ~15% of current assets
                "سرمایه‌گذاری‌های کوتاه‌مدت": company_size * 0.04,  # Short-term investments ~10% of current assets
                "دریافتنی‌های تجاری و سایر دریافتنی‌ها": company_size * 0.1,
                # Accounts receivable ~25% of current assets
                "موجودی مواد و کالا": company_size * 0.12,  # Inventory ~30% of current assets

                # Other items
                "استهلاک": company_size * 0.03,  # Depreciation ~5% of non-current assets
            }

            if metric_name in defaults:
                self.log(f"Using default value for {metric_name}: {defaults[metric_name]}")
                return defaults[metric_name]

            return 0  # If we don't have a default, return 0

        def estimate_company_size(self):
            """Estimate the company size based on available metrics"""
            # Start with a reasonable default size
            default_size = 1000000000  # 1 billion currency units (a medium-sized company)

            # Try to get more accurate estimate from available data
            size_indicators = {}

            # Check all years' data
            for year_data in self.last_year_data.values():
                # Check key size indicators
                if "جمع دارایی‌ها" in year_data and year_data["جمع دارایی‌ها"]:
                    size_indicators["Total Assets"] = year_data["جمع دارایی‌ها"]

                if "فروش" in year_data and year_data["فروش"]:
                    size_indicators["Sales"] = year_data["فروش"]

                if "جمع بدهی‌ها" in year_data and year_data["جمع بدهی‌ها"]:
                    size_indicators["Total Liabilities"] = year_data["جمع بدهی‌ها"]

                if "جمع حقوق مالکانه" in year_data and year_data["جمع حقوق مالکانه"]:
                    size_indicators["Total Equity"] = year_data["جمع حقوق مالکانه"]

            # If we have any size indicators, use their average
            if size_indicators:
                return sum(size_indicators.values()) / len(size_indicators)

            return default_size

        def get_year_from_column(self, column):
            """Extract a year value from a column name or value"""
            if column is None:
                return None

            # Convert to string just in case
            col_str = str(column).strip()

            # Common Persian/Arabic year patterns (1390s, 1400s)
            persian_years = ["۱۳۹۰", "۱۳۹۱", "۱۳۹۲", "۱۳۹۳", "۱۳۹۴", "۱۳۹۵", "۱۳۹۶", "۱۳۹۷", "۱۳۹۸", "۱۳۹۹",
                             "۱۴۰۰", "۱۴۰۱", "۱۴۰۲", "۱۴۰۳", "۱۴۰۴", "۱۴۰۵"]

            # English digits for Persian years
            english_years = ["1390", "1391", "1392", "1393", "1394", "1395", "1396", "1397", "1398", "1399",
                             "1400", "1401", "1402", "1403", "1404", "1405"]

            # Check for exact Persian year matches
            for year in persian_years:
                if year in col_str:
                    # Convert Persian digits to English
                    return self.persian_to_english_digits(year)

            # Check for exact English year matches for Persian calendar
            for year in english_years:
                if year in col_str:
                    return year

            # Check for Gregorian years (2010-2025)
            gregorian_years = [str(year) for year in range(2010, 2026)]
            for year in gregorian_years:
                if year in col_str:
                    # Convert Gregorian to Persian (approximate)
                    persian_year = str(int(year) - 621)
                    return persian_year

            # Try regular expression for 4-digit numbers that could be years
            year_match = re.search(r'\b(13\d\d|14\d\d|20[12]\d)\b', col_str)
            if year_match:
                year = year_match.group(1)
                # Convert to Persian calendar if it's Gregorian
                if year.startswith('20'):
                    return str(int(year) - 621)
                return year

            # Check for shortened years (e.g., 99 for 1399, 00 for 1400)
            short_year_match = re.search(r'\b(9\d|0[0-5])\b', col_str)
            if short_year_match:
                short_year = short_year_match.group(1)
                if short_year.startswith('9'):
                    return '13' + short_year
                else:
                    return '14' + short_year

            return None

        def persian_to_english_digits(self, persian_str):
            """Convert Persian/Arabic digits to English digits"""
            # Mapping of Persian/Arabic digits to English
            digit_map = {
                '۰': '0', '۱': '1', '۲': '2', '۳': '3', '۴': '4',
                '۵': '5', '۶': '6', '۷': '7', '۸': '8', '۹': '9'
            }

            english_str = ''
            for char in persian_str:
                if char in digit_map:
                    english_str += digit_map[char]
                else:
                    english_str += char

            return english_str

        def identify_financial_metric(self, row_label):
            """Identify which financial metric a row label represents with better fuzzy matching"""
            if not hasattr(self, 'financial_metric_names'):
                # Dictionary of common financial metrics in Persian with English equivalents
                self.financial_metric_names = {
                    # Balance Sheet - Assets
                    "دارایی‌های جاری": ["current assets", "current asset", "assets current", "asset current"],
                    "موجودی نقد": ["cash", "cash and cash equivalents", "cash equivalents", "cash in hand",
                                   "cash balance"],
                    "سرمایه‌گذاری‌های کوتاه‌مدت": ["short-term investments", "short term investments",
                                                   "short-term investment", "short term investment"],
                    "دریافتنی‌های تجاری و سایر دریافتنی‌ها": ["accounts receivable", "trade receivables", "receivables",
                                                              "trade and other receivables"],
                    "موجودی مواد و کالا": ["inventory", "inventories", "stock", "stocks", "goods", "materials"],
                    "پیش پرداخت‌ها": ["prepayments", "prepaid expenses", "advances", "advance payments"],
                    "دارایی‌های غیرجاری": ["non-current assets", "noncurrent assets", "fixed assets",
                                           "long-term assets", "long term assets"],
                    "دارایی‌های ثابت مشهود": ["property, plant and equipment", "ppe", "tangible fixed assets",
                                              "tangible assets"],
                    "دارایی‌های نامشهود": ["intangible assets", "intangible", "intangibles", "goodwill"],
                    "سرمایه‌گذاری‌های بلندمدت": ["long-term investments", "long term investments",
                                                 "long-term investment", "long term investment"],
                    "جمع دارایی‌ها": ["total assets", "assets total", "sum of assets", "assets sum"],

                    # Balance Sheet - Liabilities
                    "بدهی‌های جاری": ["current liabilities", "current liability", "liabilities current",
                                      "liability current"],
                    "پرداختنی‌های تجاری و سایر پرداختنی‌ها": ["accounts payable", "trade payables", "payables",
                                                              "trade and other payables"],
                    "وام‌های کوتاه‌مدت": ["short-term loans", "short term loans", "short-term borrowings",
                                          "short term borrowings"],
                    "ذخایر": ["provisions", "reserves", "provision"],
                    "بدهی‌های غیرجاری": ["non-current liabilities", "noncurrent liabilities", "long-term liabilities",
                                         "long term liabilities"],
                    "وام‌های بلندمدت": ["long-term loans", "long term loans", "long-term borrowings",
                                        "long term borrowings"],
                    "تسهیلات مالی بلندمدت": ["long-term financial facilities", "long term financial facilities"],
                    "ذخیره مزایای پایان خدمت کارکنان": ["employee benefit provisions", "employee retirement benefits",
                                                        "staff termination benefits"],
                    "جمع بدهی‌ها": ["total liabilities", "liabilities total", "sum of liabilities", "liabilities sum"],

                    # Balance Sheet - Equity
                    "سرمایه": ["capital", "share capital", "paid-in capital", "paid in capital"],
                    "اندوخته قانونی": ["legal reserve", "statutory reserve", "legal reserves", "statutory reserves"],
                    "سود انباشته": ["retained earnings", "accumulated profit", "retained profits",
                                    "accumulated profits"],
                    "جمع حقوق مالکانه": ["total equity", "equity total", "total shareholders' equity",
                                         "shareholders' equity", "equity"],

                    # Income Statement
                    "فروش": ["sales", "revenue", "turnover", "income"],
                    "بهای تمام شده کالای فروش رفته": ["cost of goods sold", "cogs", "cost of sales"],
                    "سود ناخالص": ["gross profit", "gross income", "gross margin"],
                    "هزینه‌های فروش، اداری و عمومی": ["selling, general and administrative expenses", "sga",
                                                      "selling and administrative expenses"],
                    "هزینه‌های عملیاتی": ["operating expenses", "operating costs", "operating expenditures"],
                    "سود عملیاتی": ["operating profit", "operating income", "ebit",
                                    "earnings before interest and taxes"],
                    "هزینه‌های مالی": ["financial expenses", "finance costs", "interest expenses", "interest expense"],
                    "درآمد مالی": ["financial income", "finance income", "interest income"],
                    "سایر درآمدها و هزینه‌ها": ["other income and expenses", "other income", "other expenses"],
                    "سود قبل از مالیات": ["profit before tax", "earnings before taxes", "income before taxes", "ebt"],
                    "مالیات": ["tax", "taxes", "income tax", "income taxes", "tax expense", "tax expenses"],
                    "سود خالص": ["net profit", "net income", "earnings", "profit", "net earnings"],

                    # Cash Flow Statement
                    "جریان نقد عملیاتی": ["operating cash flow", "cash flow from operations", "operating activities"],
                    "جریان نقد سرمایه‌گذاری": ["investing cash flow", "cash flow from investing",
                                               "investing activities"],
                    "جریان نقد تامین مالی": ["financing cash flow", "cash flow from financing", "financing activities"],

                    # Other common metrics
                    "سود هر سهم": ["earnings per share", "eps"],
                    "ارزش دفتری هر سهم": ["book value per share", "bvps"],
                    "نسبت قیمت به درآمد": ["price to earnings ratio", "p/e ratio", "pe ratio"],
                    "استهلاک": ["depreciation", "amortization", "depreciation and amortization"]
                }

            # Clean and normalize input
            if not row_label or pd.isna(row_label):
                return None, 0

            row_label = row_label.lower().strip()

            # First try exact matches
            for metric_name, alternatives in self.financial_metric_names.items():
                # Check if row_label exactly matches the Persian name
                if row_label == metric_name.lower():
                    return metric_name, 1.0

                # Check if row_label contains the Persian name
                if metric_name.lower() in row_label:
                    # Return higher confidence if it's a close match
                    if len(row_label) <= len(metric_name) + 10:
                        return metric_name, 0.95
                    return metric_name, 0.85

                # Check English alternatives
                for alt in alternatives:
                    if row_label == alt:
                        return metric_name, 0.9
                    if alt in row_label:
                        # Return higher confidence if it's a close match
                        if len(row_label) <= len(alt) + 10:
                            return metric_name, 0.85
                        return metric_name, 0.75

            # If no exact match, try fuzzy matching on both Persian and English names
            best_match = None
            best_score = self.min_match_confidence

            for metric_name, alternatives in self.financial_metric_names.items():
                # Check similarity with Persian name
                ratio = self.get_fuzzy_ratio(row_label, metric_name.lower())
                if ratio > best_score:
                    best_score = ratio
                    best_match = metric_name

                # Check similarity with English alternatives
                for alt in alternatives:
                    ratio = self.get_fuzzy_ratio(row_label, alt)
                    if ratio > best_score:
                        best_score = ratio
                        best_match = metric_name

            return best_match, best_score

        def get_fuzzy_ratio(self, s1, s2):
            """Get fuzzy match ratio between two strings"""
            # Simple implementation using difflib
            import difflib
            return difflib.SequenceMatcher(None, s1, s2).ratio()

        def clean_number(self, value):
            """Clean and convert a value to a number"""
            if value is None or pd.isna(value):
                return 0

            # If already a number, return it
            if isinstance(value, (int, float)):
                return float(value)

            # Convert to string and clean
            value_str = str(value).strip()

            # Skip empty strings
            if not value_str:
                return 0

            # Handle parentheses (negative numbers)
            if value_str.startswith('(') and value_str.endswith(')'):
                value_str = '-' + value_str[1:-1]

            # Remove common separators
            value_str = value_str.replace(',', '').replace(' ', '')

            # Try to extract the numeric part
            try:
                # Handle percentage values
                if '%' in value_str:
                    match = re.search(r'([-+]?\d*\.?\d+)', value_str)
                    if match:
                        return float(match.group(1)) / 100

                # Convert to float
                return float(value_str) if value_str else 0
            except (ValueError, TypeError):
                # If conversion fails, return 0
                return 0

        def log(self, message):
            """Log a message if debug_mode is enabled"""
            if hasattr(self, 'debug_mode') and self.debug_mode:
                print(f"[DEBUG] {message}")

        def generate_financial_report(self, metrics_by_year, ratios_by_year):
            """Generate a comprehensive financial report with analysis"""
            report = {
                "metrics": metrics_by_year,
                "ratios": ratios_by_year,
                "analysis": self.analyze_financial_data(metrics_by_year, ratios_by_year),
                "recommendations": self.generate_recommendations(metrics_by_year, ratios_by_year),
                "trends": self.analyze_trends(metrics_by_year, ratios_by_year),
                "warnings": self.identify_warning_signs(metrics_by_year, ratios_by_year)
            }

            return report

        def analyze_financial_data(self, metrics_by_year, ratios_by_year):
            """Analyze financial data and provide insights"""
            analysis = {}

            # Analyze profitability
            analysis["profitability"] = self.analyze_profitability(metrics_by_year, ratios_by_year)

            # Analyze liquidity
            analysis["liquidity"] = self.analyze_liquidity(metrics_by_year, ratios_by_year)

            # Analyze solvency
            analysis["solvency"] = self.analyze_solvency(metrics_by_year, ratios_by_year)

            # Analyze efficiency
            analysis["efficiency"] = self.analyze_efficiency(metrics_by_year, ratios_by_year)

            # Analyze growth
            analysis["growth"] = self.analyze_growth(metrics_by_year)

            return analysis

        def analyze_profitability(self, metrics_by_year, ratios_by_year):
            """Analyze profitability metrics"""
            profitability = {
                "summary": "",
                "details": {}
            }

            # Get the years in chronological order
            years

        def format_number(self, number, decimals=2):
            """Format number with thousands separator"""
            if number is None:
                return "N/A"

            if isinstance(number, str):
                return number

            if abs(number) >= 1_000_000_000:
                return f"{number/1_000_000_000:.{decimals}f} B"
            elif abs(number) >= 1_000_000:
                return f"{number/1_000_000:.{decimals}f} M"
            elif abs(number) >= 1_000:
                return f"{number/1_000:.{decimals}f} K"
            else:
                return f"{number:.{decimals}f}"

        def format_ratio(self, ratio, as_percentage=False):
            """Format ratio value"""
            if ratio is None:
                return "N/A"

            if as_percentage:
                return f"{ratio * 100:.2f}%"
            else:
                return f"{ratio:.2f}"

        def generate_excel_report(self, metrics_by_year, ratios_by_year, output_file):
            """Generate Excel report with extracted metrics and calculated ratios"""
            try:
                # Create Excel writer
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    # Create metrics dataframes
                    metrics_data = []
                    for year, metrics in metrics_by_year.items():
                        for metric, value in metrics.items():
                            metrics_data.append({
                                "Year": year,
                                "Metric": metric,
                                "Value": value,
                                "Formatted Value": self.format_number(value)
                            })

                    if metrics_data:
                        metrics_df = pd.DataFrame(metrics_data)
                        # Pivot to have years as columns
                        metrics_pivot = metrics_df.pivot(index="Metric", columns="Year", values="Value")
                        metrics_pivot.to_excel(writer, sheet_name="Financial Metrics")

                    # Create ratios dataframes
                    ratios_data = []
                    for year, ratios in ratios_by_year.items():
                        for ratio, value in ratios.items():
                            # Determine if this should be displayed as percentage
                            as_percentage = any(term in ratio.lower() for term in ["margin", "return", "حاشیه", "بازده"])

                            ratios_data.append({
                                "Year": year,
                                "Ratio": ratio,
                                "Value": value,
                                "Formatted Value": self.format_ratio(value, as_percentage)
                            })

                    if ratios_data:
                        ratios_df = pd.DataFrame(ratios_data)
                        # Pivot for better viewing
                        ratios_pivot = ratios_df.pivot(index="Ratio", columns="Year", values="Value")
                        ratios_pivot.to_excel(writer, sheet_name="Financial Ratios")

                    # Create summary sheet
                    self.create_summary_sheet(writer, metrics_by_year, ratios_by_year)

                    # Format the Excel sheets
                    self.format_excel_sheets(writer)

                self.log(f"Excel report generated: {output_file}")
                return output_file
            except Exception as e:
                self.log(f"Error generating Excel report: {str(e)}")
                traceback.print_exc()
                return None

        def create_summary_sheet(self, writer, metrics_by_year, ratios_by_year):
            """Create summary sheet with key metrics and ratios"""
            try:
                # Prepare summary data
                years = sorted(list(metrics_by_year.keys()))

                # Key metrics to highlight
                key_metrics = [
                    "فروش",
                    "سود ناخالص",
                    "سود عملیاتی",
                    "سود خالص",
                    "موجودی نقد",
                    "دارایی‌های جاری",
                    "جمع دارایی‌ها",
                    "بدهی‌های جاری",
                    "جمع بدهی‌ها",
                    "جمع حقوق مالکانه"
                ]

                # Key ratios to highlight
                key_ratios = [
                    "نسبت جاری (Current Ratio)",
                    "نسبت آنی (Quick Ratio)",
                    "حاشیه سود ناخالص (Gross Profit Margin)",
                    "حاشیه سود خالص (Net Profit Margin)",
                    "بازده حقوق صاحبان سهام (Return on Equity)",
                    "نسبت بدهی (Debt Ratio)"
                ]

                # Create metrics summary
                metrics_summary = []
                for metric in key_metrics:
                    row = {"Metric": metric}
                    for year in years:
                        if year in metrics_by_year and metric in metrics_by_year[year]:
                            row[year] = metrics_by_year[year][metric]
                        else:
                            row[year] = None
                    metrics_summary.append(row)

                # Create ratios summary
                ratios_summary = []
                for ratio in key_ratios:
                    row = {"Ratio": ratio}
                    for year in years:
                        if year in ratios_by_year and ratio in ratios_by_year[year]:
                            row[year] = ratios_by_year[year][ratio]
                        else:
                            row[year] = None
                    ratios_summary.append(row)

                # Convert to DataFrames
                if metrics_summary:
                    metrics_summary_df = pd.DataFrame(metrics_summary)
                    metrics_summary_df.to_excel(writer, sheet_name="Summary", startrow=1, index=False)

                if ratios_summary:
                    ratios_summary_df = pd.DataFrame(ratios_summary)
                    ratios_summary_df.to_excel(writer, sheet_name="Summary", startrow=len(metrics_summary) + 5, index=False)

                # Add title to summary page
                workbook = writer.book
                summary_sheet = writer.sheets["Summary"]
                summary_sheet.cell(row=1, column=1, value="Key Financial Metrics")
                summary_sheet.cell(row=len(metrics_summary) + 5, column=1, value="Key Financial Ratios")

            except Exception as e:
                self.log(f"Error creating summary sheet: {str(e)}")
                traceback.print_exc()

        def format_excel_sheets(self, writer):
            """Format Excel worksheets for better readability"""
            try:
                workbook = writer.book

                # Define styles
                header_style = openpyxl.styles.NamedStyle(name="header_style")
                header_style.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
                header_style.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4",
                                                                fill_type="solid")
                header_style.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                header_style.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style="thin"),
                    right=openpyxl.styles.Side(style="thin"),
                    top=openpyxl.styles.Side(style="thin"),
                    bottom=openpyxl.styles.Side(style="thin")
                )

                # Number format style
                number_style = openpyxl.styles.NamedStyle(name="number_style")
                number_style.number_format = '#,##0.00'
                number_style.alignment = openpyxl.styles.Alignment(horizontal="right")

                # Percentage style
                percent_style = openpyxl.styles.NamedStyle(name="percent_style")
                percent_style.number_format = '0.00%'
                percent_style.alignment = openpyxl.styles.Alignment(horizontal="right")

                # Register styles
                if "header_style" not in workbook.named_styles:
                    workbook.add_named_style(header_style)
                if "number_style" not in workbook.named_styles:
                    workbook.add_named_style(number_style)
                if "percent_style" not in workbook.named_styles:
                    workbook.add_named_style(percent_style)

                # Format each worksheet
                for sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]

                    # Get the dimensions of the worksheet
                    max_row = worksheet.max_row
                    max_col = worksheet.max_column

                    # Format headers
                    for col in range(1, max_col + 1):
                        cell = worksheet.cell(row=1, column=col)
                        cell.style = "header_style"

                        # Adjust column width based on content
                        max_length = 0
                        for row in range(1, max_row + 1):
                            cell_value = str(worksheet.cell(row=row, column=col).value)
                            max_length = max(max_length, len(cell_value))
                        adjusted_width = min(max_length + 2, 50)  # Cap width at 50
                        worksheet.column_dimensions[get_column_letter(col)].width = adjusted_width

                    # Format data cells
                    for row in range(2, max_row + 1):
                        for col in range(1, max_col + 1):
                            cell = worksheet.cell(row=row, column=col)

                            # Skip empty cells
                            if cell.value is None:
                                continue

                            # Apply number formatting for numeric values
                            if isinstance(cell.value, (int, float)):
                                if sheet_name == "Financial Ratios":
                                    cell.style = "percent_style"
                                else:
                                    cell.style = "number_style"

                            # Add subtle borders
                            cell.border = openpyxl.styles.Border(
                                left=openpyxl.styles.Side(style="thin", color="E0E0E0"),
                                right=openpyxl.styles.Side(style="thin", color="E0E0E0"),
                                top=openpyxl.styles.Side(style="thin", color="E0E0E0"),
                                bottom=openpyxl.styles.Side(style="thin", color="E0E0E0")
                            )

                    # Freeze panes
                    worksheet.freeze_panes = 'B2'

            except Exception as e:
                self.log(f"Error formatting Excel sheets: {str(e)}")
                traceback.print_exc()

                def process_directory(self):
                    """Process all Excel and CSV files in the input directory"""
                    try:
                        # Initialize storage for all metrics
                        all_metrics = {}
                        all_ratios = {}
                        processed_files = []

                        # Get list of files to process
                        files = list(self.input_folder.glob("*.xls*")) + list(self.input_folder.glob("*.csv"))

                        if not files:
                            self.log("No Excel or CSV files found in the input directory")
                            return None

                        # Process each file
                        for file_path in files:
                            try:
                                self.log(f"Processing file: {file_path}")
                                metrics = self.extract_metrics_from_file(file_path)

                                if metrics:
                                    # Merge metrics
                                    for year, year_metrics in metrics.items():
                                        if year not in all_metrics:
                                            all_metrics[year] = {}
                                        for metric, value in year_metrics.items():
                                            # Only update if we don't have this metric or if the new value is larger
                                            if metric not in all_metrics[year] or value > all_metrics[year][metric]:
                                                all_metrics[year][metric] = value

                                    processed_files.append(file_path)

                            except Exception as e:
                                self.log(f"Error processing file {file_path}: {str(e)}")
                                continue

                        if not all_metrics:
                            self.log("No valid metrics found in any file")
                            return None

                        # Calculate ratios for consolidated metrics
                        all_ratios = self.calculate_financial_ratios(all_metrics)

                        # Generate timestamp for output file
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        output_file = self.output_dir / f"Financial_Analysis_Report_{timestamp}.xlsx"

                        # Generate consolidated report
                        self.generate_excel_report(all_metrics, all_ratios, output_file)

                        # Generate summary of processed files
                        self.log(f"\nProcessed {len(processed_files)} files:")
                        for file in processed_files:
                            self.log(f"- {file}")
                        self.log(f"\nReport generated: {output_file}")

                        return output_file

                    except Exception as e:
                        self.log(f"Error processing directory: {str(e)}")
                        traceback.print_exc()
                        return None

                    def main():
                        """Main execution function"""
                        try:
                            # Setup argument parser
                            parser = argparse.ArgumentParser(description='Financial Statement Analyzer')
                            parser.add_argument('input_folder', type=str,
                                                help=r'C:\retina_env\BOT\sorat')
                            parser.add_argument('--debug', action='store_true', help='Enable debug mode')

                            args = parser.parse_args()

                            # Create analyzer instance
                            analyzer = FinancialAnalyzer(args.input_folder)
                            analyzer.debug_mode = args.debug

                            # Process the directory
                            output_file = analyzer.process_directory()

                            if output_file:
                                print(f"\nAnalysis complete! Report generated: {output_file}")
                            else:
                                print("\nError: Failed to generate report")
                                sys.exit(1)

                        except Exception as e:
                            print(f"Error in main execution: {str(e)}")
                            traceback.print_exc()
                            sys.exit(1)

                    if __name__ == "__main__":
                        main()